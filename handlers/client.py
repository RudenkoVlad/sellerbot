import os
import types

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import state

from states.client_states import SearchPrice, DeleteItemFromCart, SearchName
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from create_bot import dp, bot
from keyboards.admin_kb import kb_admin, item_btn_admin
from keyboards.client_kb import kb_client, item_btn_client
from keyboards.catalog_kb import create_keyboard_catalog
from keyboards.cart_kb import cart_btn
from keyboards.info_kb import info_change_btn
from keyboards.view_products_kb import products_kb

from database import database as db

save_id = None


@dp.message_handler(commands=['start'])
async def cmd_start(message: types.Message):
    global save_id
    save_id = message.from_user.id

    await db.add_user(message.from_user.id)

    if message.from_user.id == int(os.getenv('ADMIN_ID')):
        await message.answer('Вітаю, Ви увійшли як адмін!', reply_markup=kb_admin)
    else:
        await message.answer(
            f"Привіт, {message.from_user.first_name}! Вітаю в нашому магазині.",
            reply_markup=kb_client)


# ----------------------------------------------------------------------------------------------------------------------
@dp.message_handler(text='Переглянути товари')
async def view_products(message: types.Message):
    await message.answer('Оберіть з клавіатури, що ви хочете зробити (підбір товарів по ціні, пошук за назвою товара '
                         'або просто перегляньте каталог)', reply_markup=products_kb)


@dp.message_handler(text='Повернутися назад')
async def back_to_start_kb(message: types.Message):
    if message.from_user.id == int(os.getenv('ADMIN_ID')):
        keyboard = kb_admin
    else:
        keyboard = kb_client

    await message.answer('Повернення до головної клавіатури', reply_markup=keyboard)


# ----------------------------------------------------------------------------------------------------------------------
@dp.message_handler(text='Підбір товарів по ціні')
async def search_by_price(message: types.Message):
    await SearchPrice.waiting_category.set()
    await show_catalog(message)


@dp.callback_query_handler(lambda query: query.data.isdigit(), state=SearchPrice.waiting_category)
async def handle_category_selection(callback_query: types.CallbackQuery, state: FSMContext):
    category_id = int(callback_query.data)
    await state.update_data(category_id=category_id)
    await ask_price_range(callback_query.message)


async def ask_price_range(message: types.Message):
    await message.answer("Введіть ціновий діапазон у форматі 'мінімальна ціна - максимальна ціна', наприклад, 10-50:")
    await SearchPrice.waiting_price_range.set()

min_max = {
    'min': None,
    'max': None,
    'category_id': None
}


@dp.message_handler(state=SearchPrice.waiting_price_range)
async def handle_price_range_input(message: types.Message, state: FSMContext):
    global min_max
    current_category_messages.clear()
    price_range = message.text.strip()
    min_price, max_price = map(int, price_range.split('-'))
    min_max['min'] = min_price
    min_max['max'] = max_price
    async with state.proxy() as data:
        category_id = data.get('category_id')
        min_max['category_id'] = category_id

        items = await db.get_items_by_price_range_and_category(min_price, max_price, category_id)

        if items:
            await show_item(message, items[0])
            await state.finish()
        else:
            await message.answer("Товарів у заданому ціновому діапазоні не знайдено.")


# ----------------------------------------------------------------------------------------------------------------------
@dp.message_handler(text='Пошук по назві')
async def search_by_name(message: types.Message):
    await SearchName.waiting_category.set()
    await show_catalog(message)


@dp.callback_query_handler(lambda query: query.data.isdigit(), state=SearchName.waiting_category)
async def handle_category_selection_name(callback_query: types.CallbackQuery, state: FSMContext):
    category_id = int(callback_query.data)
    await state.update_data(category_id=category_id)
    await ask_name(callback_query.message)


async def ask_name(message: types.Message):
    await message.answer("Введіть назву товару або її частину:")
    await SearchName.waiting_name.set()


@dp.message_handler(state=SearchName.waiting_name)
async def handle_name_input(message: types.Message, state: FSMContext):
    current_category_messages.clear()
    name = message.text.strip()
    async with state.proxy() as data:
        category_id = data.get('category_id')

        items = await db.search_items_by_name(category_id, name)
        print(items)

        if items:
            await show_item(message, items[0])
            await state.finish()
        else:
            await message.answer("Товарів із заданою назвою не знайдено.")
            # await state.finish()


# ----------------------------------------------------------------------------------------------------------------------
current_category_messages = {}


@dp.message_handler(text='Каталог')
async def catalog(message: types.Message):
    global current_category_messages
    current_category_messages = {}

    for key in min_max:
        min_max[key] = None

    await show_catalog(message)


async def show_catalog(message: types.Message):
    categories = await db.get_all_categories()

    if categories:
        keyboard = await create_keyboard_catalog(categories)
        await message.answer('Виберіть категорію, щоб побачити товар: ', reply_markup=keyboard)
    else:
        await message.answer('Немає доступних категорій')


item_data = {
    "item_id": None,
    "category_id": None
}


async def show_item(message: types.Message, item):
    global item_data
    item_data["item_id"] = item[0]
    item_data["category_id"] = item[1]

    item_info = f"Name: {item[2]}\n"
    item_info += f"Description: {item[3]}\n"
    item_info += f"Price: {item[4]}\n"

    keyboard = item_btn_admin if save_id == int(os.getenv('ADMIN_ID')) else item_btn_client

    if message.chat.id in current_category_messages:
        await bot.edit_message_media(chat_id=message.chat.id, message_id=current_category_messages[message.chat.id],
                                     media=types.InputMediaPhoto(media=item[5], caption=item_info),
                                     reply_markup=keyboard)
    else:
        sent_message = await bot.send_photo(chat_id=message.chat.id, photo=item[5], caption=item_info,
                                            reply_markup=keyboard)
        current_category_messages[message.chat.id] = sent_message.message_id


async def show_items_in_categories(message: types.Message, category: str):
    items = await db.get_items_by_category(category)

    if items:
        await show_item(message, items[0])
    else:
        if message.chat.id in current_category_messages:
            await bot.delete_message(message.chat.id,
                                     current_category_messages[message.chat.id])
            del current_category_messages[message.chat.id]
        # await message.answer('В цій категорії товарів не знайдено')


@dp.callback_query_handler(lambda query: query.data.isdigit())
async def show_category_items(callback_query: types.CallbackQuery):
    category = callback_query.data
    await show_items_in_categories(callback_query.message, category)


@dp.callback_query_handler(text=['previous_item', 'next_item'])
async def navigate_items(callback_query: types.CallbackQuery):

    if min_max['min'] is not None and min_max['max'] is not None:
        items = await db.get_items_by_price_range_and_category(min_max['min'], min_max['max'], min_max['category_id'])
    else:
        items = await db.get_items_by_category(item_data["category_id"])

    current_index = next((index for index, item in enumerate(items) if item[0] == item_data["item_id"]), None)

    if current_index is not None:
        if callback_query.data == 'previous_item' and current_index > 0:
            await show_item(callback_query.message, items[current_index - 1])
        elif callback_query.data == 'next_item' and current_index < len(items) - 1:
            await show_item(callback_query.message, items[current_index + 1])


@dp.callback_query_handler(text='add_to_cart')
async def add_item_to_cart(callback_query: types.CallbackQuery):
    cart_items = await db.get_items_in_cart(save_id)

    if item_data["item_id"] in cart_items:
        await callback_query.answer('Товар вже знаходиться в кошику')
    else:
        await db.add_to_cart(save_id, item_data["item_id"])
        await callback_query.answer('Товар додано до кошика')


# ----------------------------------------------------------------------------------------------------------------------
summa = 0


async def display_basket(message: types.Message):
    global summa
    summa = 0
    cart_items = await db.get_items_in_cart(save_id)

    if not cart_items:
        await message.answer('Кошик пустий!')
    else:
        response = "🛒 Ваш кошик:\n"
        for item_id in cart_items:
            item = await db.get_item(item_id)
            response += f"ID: {item_id}\nНазва: {item[2]}\nЦіна: {item[4]}\n\n"
            summa += int(item[4])
        response += f"💰 Всього: {summa}"
        await message.answer(response, reply_markup=cart_btn)


@dp.message_handler(text='Кошик')
async def basket(message: types.Message):
    await display_basket(message)


@dp.callback_query_handler(text='clear_cart')
async def clear_cart(callback_query: types.CallbackQuery):
    await db.delete_items_from_cart(save_id)
    await callback_query.answer('Кошик очищено')

    cart_item = await db.get_items_in_cart(save_id)
    if not cart_item:
        await callback_query.message.delete()
        await callback_query.message.answer('Ваш кошик пустий')


@dp.callback_query_handler(text='delete_item_from_cart')
async def delete_item_query(callback_query: types.CallbackQuery, state: FSMContext):
    await callback_query.message.answer('Введіть ID товару, який ви хочете видалити з кошика:')
    await DeleteItemFromCart.item_id.set()


@dp.message_handler(state=DeleteItemFromCart.item_id)
async def delete_item(message: types.Message, state: FSMContext):
    item_id = int(message.text)
    await db.delete_item_from_cart(save_id, item_id)
    await message.answer(f'Товар з ID {item_id} був видалений з кошика')
    await state.finish()
    await display_basket(message)


# ----------------------------------------------------------------------------------------------------------------------


@dp.callback_query_handler(text='to_pay')
async def to_pay(callback_query: types.CallbackQuery):
    await callback_query.message.answer('Для оплати товарів натисніть "Заплатити" і введіть всю потрібну інформацію.'
                                        'Будь-ласка заповнюйте поля коректними данними. '
                                        'В поле вулиці можете вводить адрес пошти. Якщо виникнуть непорозуміння щодо '
                                        'заповнених вами даних, то наша адміністрація з вами зв\'яжиться для уточнення')
    await bot.send_invoice(callback_query.message.chat.id, 'Покупка товарів', 'Покупка товарів в телеграм-магазині',
                           'invoice', os.getenv('PAYMENT_TOKEN'), 'USD',
                           [types.LabeledPrice('Покупка товарів', summa * 100)],
                           need_email=True, need_phone_number=True, need_shipping_address=True)


@dp.pre_checkout_query_handler(lambda query: True)
async def pre_checkout_query(pre_checkout_q: types.PreCheckoutQuery):
    await bot.answer_pre_checkout_query(pre_checkout_q.id, ok=True)


@dp.message_handler(content_types=types.ContentType.SUCCESSFUL_PAYMENT)
async def successful_payment(message: types.Message):
    payment_info = message.successful_payment.to_python()

    # Dividing total_amount by 100
    payment_info['total_amount'] = payment_info.get('total_amount', 0) / 100

    # Load existing data from the file if it exists
    try:
        workbook = load_workbook('payment_info.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook and select the active worksheet
        workbook = Workbook()
        sheet = workbook.active
        # Write headers
        sheet.append(
            ['currency', 'total_amount', 'ordered_items', 'invoice_payload', 'phone_number', 'email', 'country_code',
             'state', 'city', 'street_line1', 'street_line2', 'post_code', 'telegram_payment_charge_id',
             'provider_payment_charge_id'])

    # Extract ordered items
    ordered_items = ''
    cart_items = await db.get_items_in_cart(save_id)
    for item_id in cart_items:
        item = await db.get_item(item_id)
        ordered_items += f"{item[2]} ({item[4]}), "

    # Write payment information excluding order_info
    order_info = payment_info.pop('order_info')
    order_info_values = [order_info.get('phone_number', ''),
                         order_info.get('email', ''),
                         order_info.get('shipping_address', {}).get('country_code', ''),
                         order_info.get('shipping_address', {}).get('state', ''),
                         order_info.get('shipping_address', {}).get('city', ''),
                         order_info.get('shipping_address', {}).get('street_line1', ''),
                         order_info.get('shipping_address', {}).get('street_line2', ''),
                         order_info.get('shipping_address', {}).get('post_code', '')]

    sheet.append([payment_info.get('currency', ''),
                  payment_info.get('total_amount', ''),
                  ordered_items.rstrip(', '),  # Remove trailing comma and space
                  payment_info.get('invoice_payload', '')] + order_info_values +
                 [payment_info.get('telegram_payment_charge_id', ''),
                  payment_info.get('provider_payment_charge_id', '')])

    # Save the workbook
    workbook.save('payment_info.xlsx')

    await bot.send_message(message.chat.id,
                           f"Платіж на суму {message.successful_payment.total_amount // 100} "
                           f"{message.successful_payment.currency} пройшов успішно! Очікуйте ваш товар.")


# ----------------------------------------------------------------------------------------------------------------------
info_text = "Мій творець: @BJLAD_IK"


@dp.message_handler(text='Інфо')
async def info(message: types.Message):
    global info_text

    if message.from_user.id == int(os.getenv('ADMIN_ID')):
        await message.answer(info_text, reply_markup=info_change_btn)
    else:
        await message.answer(info_text)
